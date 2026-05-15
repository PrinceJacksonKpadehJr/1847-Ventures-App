from __future__ import annotations

from datetime import date
from pathlib import Path
import random

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def build_farmer_profiles() -> list[dict[str, object]]:
    random.seed(1847)

    first_names = [
        "Kwame", "Ama", "Kojo", "Efua", "Yaw", "Adwoa", "Kofi", "Akosua", "Nana", "Abena",
        "Mensah", "Ophelia", "Yawson", "Mabel", "Kwaku", "Akua", "Fiifi", "Afi", "Kojoa", "Esi",
        "Daniel", "Patience", "Richmond", "Vida", "Samuel", "Grace", "Michael", "Ruth", "Joseph", "Janet",
        "Emmanuel", "Priscilla", "Theophilus", "Doris", "Augustine", "Martha",
    ]
    last_names = [
        "Asare", "Boateng", "Mensah", "Owusu", "Addo", "Nyarko", "Ofori", "Ampofo", "Agyeman", "Amoah",
        "Boadu", "Adjei", "Annan", "Sarpong", "Tetteh", "Darko", "Arthur", "Kusi", "Baah", "Oppong",
    ]
    regions = ["Ashanti", "Western North", "Bono East", "Eastern", "Ahafo", "Central"]
    districts = {
        "Ashanti": ["Bekwai", "Mampong", "Obuasi"],
        "Western North": ["Sefwi Wiawso", "Bibiani", "Bodi"],
        "Bono East": ["Techiman", "Nkoranza", "Atebubu"],
        "Eastern": ["Nkawkaw", "Koforidua", "Akim Oda"],
        "Ahafo": ["Goaso", "Kenyasi", "Hwidiem"],
        "Central": ["Assin Fosu", "Twifo Praso", "Kasoa"],
    }

    rows: list[dict[str, object]] = []
    for idx in range(1, 37):
        region = random.choice(regions)
        district = random.choice(districts[region])
        community = f"{district} Community {random.randint(1, 12)}"
        first = first_names[idx - 1]
        last = random.choice(last_names)
        farmer_name = f"{first} {last}"
        farmer_id = f"FMR-{2026}-{idx:03d}"
        farm_name = f"{last} Cocoa Farm {idx:02d}"

        size_ha = round(random.uniform(1.2, 9.5), 2)
        cocoa_pct = random.randint(55, 90)
        shade_tree_pct = random.randint(10, 45)
        food_crop_pct = max(0, 100 - cocoa_pct)
        agroforestry = "yes" if shade_tree_pct >= 24 else "no"
        irrigation = random.choice(["none", "rainfed", "drip"])
        mechanization = random.choice(["manual", "semi_mechanized"])
        certification = random.choice(["Certified", "Pending", "None"])

        rows.append(
            {
                "farmer_id": farmer_id,
                "farmer_name": farmer_name,
                "farm_name": farm_name,
                "region": region,
                "district": district,
                "community": community,
                "size_in_hectares": size_ha,
                "cocoa_composition_pct": cocoa_pct,
                "food_crop_composition_pct": food_crop_pct,
                "shade_tree_cover_pct": shade_tree_pct,
                "agroforestry": agroforestry,
                "irrigation_type": irrigation,
                "mechanization_level": mechanization,
                "certification_status": certification,
                "cooperative_member": random.choice(["yes", "no"]),
                "primary_soil_type": random.choice(["loamy", "clay loam", "sandy loam"]),
                "gps_latitude": round(5.2 + random.random() * 2.9, 6),
                "gps_longitude": round(-2.9 + random.random() * 2.1, 6),
                "created_year": random.randint(2015, 2024),
            }
        )
    return rows


def build_annual_rows(farmers: list[dict[str, object]]) -> list[dict[str, object]]:
    random.seed(4781)
    rows: list[dict[str, object]] = []

    for idx, farmer in enumerate(farmers):
        if idx < 20:
            start_year = random.randint(2018, 2020)
            end_year = 2025
        else:
            start_year = random.randint(2021, 2024)
            end_year = 2025

        size_ha = float(farmer["size_in_hectares"])
        base_yield_kg_ha = random.uniform(260, 820)

        for year in range(start_year, end_year + 1):
            weather_factor = random.uniform(0.82, 1.14)
            training_bonus = 1.05 if farmer["cooperative_member"] == "yes" else 1.0
            agro_bonus = 1.06 if farmer["agroforestry"] == "yes" else 0.96
            yield_per_ha = clamp(base_yield_kg_ha * weather_factor * training_bonus * agro_bonus, 180, 980)
            yield_kg = round(yield_per_ha * size_ha, 1)

            cocoa_price = random.uniform(2.35, 3.2)
            revenue_usd = round(yield_kg * cocoa_price, 2)
            operating_cost = round((size_ha * random.uniform(320, 520)) + random.uniform(120, 580), 2)
            net_income = round(revenue_usd - operating_cost, 2)

            fertilizer_bags = random.randint(0, 7)
            emissions = round((yield_kg * random.uniform(0.17, 0.43)) + (fertilizer_bags * 75), 2)
            removals = round((float(farmer["shade_tree_cover_pct"]) * random.uniform(3.0, 6.2)), 2)
            carbon_balance = round(emissions - removals, 2)

            rows.append(
                {
                    "farmer_id": farmer["farmer_id"],
                    "farmer_name": farmer["farmer_name"],
                    "farm_name": farmer["farm_name"],
                    "region": farmer["region"],
                    "district": farmer["district"],
                    "community": farmer["community"],
                    "year": year,
                    "harvest_date": date(year, random.randint(9, 12), random.randint(1, 28)).isoformat(),
                    "size_in_hectares": size_ha,
                    "yield_kg": yield_kg,
                    "yield_per_hectare": round(yield_per_ha, 2),
                    "revenue_usd": revenue_usd,
                    "operating_cost_usd": operating_cost,
                    "net_income_usd": net_income,
                    "fertilizer_bags": fertilizer_bags,
                    "emissions_co2e_kg": emissions,
                    "carbon_removals_co2e_kg": removals,
                    "carbon_balance_co2e_kg": carbon_balance,
                    "traceability_score": random.randint(40, 98),
                    "risk_score": random.randint(18, 86),
                    "esg_score": random.randint(35, 95),
                    "verification_state": random.choice(["Verified", "Review"]),
                    "photos_complete": random.choice(["yes", "no"]),
                    "burns_farm_waste": random.choice(["never", "sometimes", "often"]),
                    "has_shade_trees": "many" if int(farmer["shade_tree_cover_pct"]) >= 32 else "few",
                    "agroforestry": farmer["agroforestry"],
                    "certification_status": farmer["certification_status"],
                }
            )

    return rows


def write_sheet(workbook: Workbook, title: str, rows: list[dict[str, object]]) -> None:
    ws = workbook.create_sheet(title=title)
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, start=1):
        width = max(14, min(36, len(header) + 3))
        ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> None:
    output_path = Path("media/investor_uploads/2026/partner_dashboard_farmer_test_dataset_2026.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    farmers = build_farmer_profiles()
    annual = build_annual_rows(farmers)

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(workbook, "farmers_master", farmers)
    write_sheet(workbook, "annual_performance", annual)

    summary = workbook.create_sheet(title="README")
    summary.append(["Dataset", "Partner Dashboard Farmer Test Dataset"])
    summary.append(["Generated On", date.today().isoformat()])
    summary.append(["Farmers", len(farmers)])
    summary.append(["Annual Records", len(annual)])
    summary.append(["Notes", "Includes 30+ farmers, varied regions, and multi-year records (some >5 years)."])

    workbook.save(output_path)
    print(f"Created: {output_path}")
    print(f"Farmers: {len(farmers)}")
    print(f"Annual rows: {len(annual)}")


if __name__ == "__main__":
    main()
