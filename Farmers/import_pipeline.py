import csv
import io
import json
import math
from collections import Counter
from datetime import date, datetime
from decimal import Decimal
from typing import Any


MISSING_TOKENS = {"", "na", "n/a", "null", "none", "nan", "-", "--"}
DATE_FORMATS = [
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%m-%d-%Y",
    "%d %b %Y",
    "%d %B %Y",
]

COLUMNAR_SAMPLE_LIMIT = 5000
CATEGORICAL_TOP_LIMIT = 20
FARM_LIFECYCLE_YOUNG_MAX = 5
FARM_LIFECYCLE_PEAK_MIN = 10
FARM_LIFECYCLE_PEAK_MAX = 20
FARM_LIFECYCLE_DECLINE_MIN = 25


class DatasetImportError(Exception):
    pass


def detect_file_format(filename: str) -> str:
    lower_name = (filename or "").lower()
    if lower_name.endswith(".csv"):
        return "csv"
    if lower_name.endswith(".xlsx"):
        return "xlsx"
    if lower_name.endswith(".json"):
        return "json"
    raise DatasetImportError("Unsupported file type. Upload .csv, .xlsx, or .json files only.")


def _normalize_header(value: Any, index: int) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return f"column_{index + 1}"
    normalized = "_".join(raw.replace("-", " ").split())
    return normalized or f"column_{index + 1}"


def _normalize_missing(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text.lower() in MISSING_TOKENS:
            return None
        return text
    return value


def _parse_boolean(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "yes", "y"}:
            return True
        if lowered in {"false", "no", "n"}:
            return False
    return None


def _parse_integer(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, Decimal) and value == int(value):
        return int(value)
    if isinstance(value, str):
        token = value.strip().replace(",", "")
        if token.startswith("$"):
            token = token[1:]
        if token.lstrip("+-").isdigit():
            try:
                return int(token)
            except ValueError:
                return None
    return None


def _parse_float(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    if isinstance(value, str):
        token = value.strip().replace(",", "")
        if token.startswith("$"):
            token = token[1:]
        try:
            return float(token)
        except ValueError:
            return None
    return None


def _parse_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        token = value.strip()
        if not token:
            return None
        try:
            return datetime.fromisoformat(token).date().isoformat()
        except ValueError:
            pass
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(token, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _value_kind(value: Any) -> str:
    if value is None:
        return "missing"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int,)):
        return "integer"
    if isinstance(value, (float, Decimal)):
        return "float"
    if isinstance(value, (datetime, date)):
        return "date"

    if _parse_boolean(value) is not None:
        return "boolean"
    if _parse_integer(value) is not None:
        return "integer"
    if _parse_float(value) is not None:
        return "float"
    if _parse_date(value) is not None:
        return "date"
    return "string"


def _final_column_type(kind_counts: Counter) -> str:
    non_missing_total = sum(
        count
        for kind, count in kind_counts.items()
        if kind != "missing"
    )
    if non_missing_total == 0:
        return "string"

    boolean_ratio = kind_counts.get("boolean", 0) / non_missing_total
    integer_ratio = kind_counts.get("integer", 0) / non_missing_total
    float_ratio = kind_counts.get("float", 0) / non_missing_total
    numeric_ratio = (kind_counts.get("integer", 0) + kind_counts.get("float", 0)) / non_missing_total
    date_ratio = kind_counts.get("date", 0) / non_missing_total

    # Use dominant-type inference so a few noisy cells do not collapse the column to string.
    if boolean_ratio >= 0.95:
        return "boolean"
    if integer_ratio == 1:
        return "integer"
    if numeric_ratio >= 0.7:
        return "float" if float_ratio > 0 else "integer"
    if date_ratio >= 0.7:
        return "date"
    return "string"


def _coerce_value(value: Any, target_type: str) -> tuple[Any, bool]:
    if value is None:
        return None, False

    if target_type == "boolean":
        parsed = _parse_boolean(value)
        if parsed is None:
            return str(value).strip(), True
        return parsed, False

    if target_type == "integer":
        parsed = _parse_integer(value)
        if parsed is None:
            return str(value).strip(), True
        return parsed, False

    if target_type == "float":
        parsed = _parse_float(value)
        if parsed is None:
            return str(value).strip(), True
        return parsed, False

    if target_type == "date":
        parsed = _parse_date(value)
        if parsed is None:
            return str(value).strip(), True
        return parsed, False

    return str(value).strip(), False


def _fingerprint(row: dict[str, Any]) -> str:
    return json.dumps(row, sort_keys=True, default=str)


def _read_csv(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            decoded = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if decoded is None:
        raise DatasetImportError("Could not decode CSV file.")

    reader = csv.reader(io.StringIO(decoded))
    rows = list(reader)
    if not rows:
        raise DatasetImportError("The uploaded file is empty.")

    headers = [_normalize_header(item, idx) for idx, item in enumerate(rows[0])]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(cell.strip() for cell in row if isinstance(cell, str)) and not any(
            cell for cell in row if not isinstance(cell, str)
        ):
            continue
        values = list(row) + [None] * (len(headers) - len(row))
        record = {headers[idx]: _normalize_missing(values[idx]) for idx in range(len(headers))}
        records.append(record)

    return headers, records


def _read_xlsx(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DatasetImportError("XLSX support requires openpyxl. Install it with pip install openpyxl.") from exc

    workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)

    try:
        first_row = next(iterator)
    except StopIteration as exc:
        raise DatasetImportError("The uploaded file is empty.") from exc

    headers = [_normalize_header(item, idx) for idx, item in enumerate(first_row)]
    records: list[dict[str, Any]] = []
    for row in iterator:
        if not any(item not in (None, "") for item in row):
            continue
        values = list(row) + [None] * (len(headers) - len(row))
        record = {headers[idx]: _normalize_missing(values[idx]) for idx in range(len(headers))}
        records.append(record)

    return headers, records


def _read_json(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        decoded = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise DatasetImportError("Could not decode JSON file.") from exc

    try:
        payload = json.loads(decoded)
    except json.JSONDecodeError as exc:
        raise DatasetImportError("Invalid JSON file.") from exc

    if isinstance(payload, dict):
        if isinstance(payload.get("rows"), list):
            raw_rows = payload.get("rows")
        elif isinstance(payload.get("data"), list):
            raw_rows = payload.get("data")
        else:
            raise DatasetImportError("JSON object must contain a 'rows' or 'data' array.")
    elif isinstance(payload, list):
        raw_rows = payload
    else:
        raise DatasetImportError("JSON payload must be an array of records.")

    records: list[dict[str, Any]] = []
    all_keys: list[str] = []
    seen = set()
    for entry in raw_rows:
        if not isinstance(entry, dict):
            continue
        for key in entry.keys():
            normalized = _normalize_header(key, len(all_keys))
            if normalized not in seen:
                seen.add(normalized)
                all_keys.append(normalized)

    if not all_keys:
        raise DatasetImportError("No tabular record rows found in JSON payload.")

    for entry in raw_rows:
        if not isinstance(entry, dict):
            continue
        normalized_entry = {
            _normalize_header(key, idx): value
            for idx, (key, value) in enumerate(entry.items())
        }
        row = {}
        for header in all_keys:
            value = normalized_entry.get(header)
            row[header] = _normalize_missing(value)
        records.append(row)

    return all_keys, records


def _detect_geospatial_fields(headers: list[str]) -> list[str]:
    geo_tokens = (
        "gps",
        "latitude",
        "longitude",
        "lat",
        "lng",
        "geohash",
        "location",
        "district",
        "region",
        "community",
        "shape",
        "polygon",
    )
    detected = []
    for header in headers:
        token = header.lower()
        if any(item in token for item in geo_tokens):
            detected.append(header)
    return detected


def _detect_lifecycle_indicators(headers: list[str]) -> list[str]:
    lifecycle_tokens = (
        "tree_age",
        "cocoa_tree_age",
        "planting_date",
        "expected_harvest_date",
        "current_stage",
        "yield_estimate",
    )
    result = []
    for header in headers:
        lowered = header.lower()
        if any(token in lowered for token in lifecycle_tokens):
            result.append(header)
    return result


def _to_float(value: Any) -> float | None:
    parsed = _parse_float(value)
    return parsed if parsed is not None else None


def _find_column(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    lowered = {header.lower(): header for header in headers}
    for candidate in candidates:
        if candidate in lowered:
            return lowered[candidate]
    for header in headers:
        token = header.lower()
        if any(candidate in token for candidate in candidates):
            return header
    return None


def _estimate_cocoa_lifecycle(rows: list[dict[str, Any]], headers: list[str]) -> dict[str, Any]:
    age_col = _find_column(headers, ("cocoa_tree_age", "tree_age", "farm_age", "age_years"))
    planting_date_col = _find_column(headers, ("planting_date", "date_planted"))
    stage_col = _find_column(headers, ("current_stage", "growth_stage"))

    if not rows:
        return {
            "estimated_stage": "unknown",
            "aging_farms": 0,
            "decline_risk_level": "medium",
            "replanting_need_ratio": 0.0,
            "long_term_roi_potential": "moderate",
        }

    ages: list[float] = []
    today = datetime.utcnow().date()
    for row in rows:
        age_value = _to_float(row.get(age_col)) if age_col else None
        if age_value is not None:
            ages.append(age_value)
            continue
        if planting_date_col:
            parsed = _parse_date(row.get(planting_date_col))
            if parsed:
                try:
                    planted = datetime.fromisoformat(parsed).date()
                    age_years = max(0, (today - planted).days / 365.25)
                    ages.append(round(age_years, 2))
                except ValueError:
                    continue

    if not ages and stage_col:
        stage_tokens = Counter()
        for row in rows:
            stage_tokens[str(row.get(stage_col) or "").strip().lower()] += 1
        common = stage_tokens.most_common(1)
        mapped = "unknown"
        if common:
            token = common[0][0]
            if any(item in token for item in ("seed", "nursery", "young", "juvenile")):
                mapped = "early_growth"
            elif any(item in token for item in ("peak", "productive", "mature")):
                mapped = "peak_productivity"
            elif any(item in token for item in ("old", "late", "decline", "aging")):
                mapped = "declining"
        return {
            "estimated_stage": mapped,
            "aging_farms": 0,
            "decline_risk_level": "medium",
            "replanting_need_ratio": 0.0,
            "long_term_roi_potential": "moderate",
        }

    if not ages:
        return {
            "estimated_stage": "unknown",
            "aging_farms": 0,
            "decline_risk_level": "medium",
            "replanting_need_ratio": 0.0,
            "long_term_roi_potential": "moderate",
        }

    aging_count = sum(1 for value in ages if value >= FARM_LIFECYCLE_DECLINE_MIN)
    decline_ratio = aging_count / len(ages)
    avg_age = sum(ages) / len(ages)

    if avg_age <= FARM_LIFECYCLE_YOUNG_MAX:
        stage = "early_growth"
        roi = "emerging"
    elif FARM_LIFECYCLE_PEAK_MIN <= avg_age <= FARM_LIFECYCLE_PEAK_MAX:
        stage = "peak_productivity"
        roi = "high"
    elif avg_age >= FARM_LIFECYCLE_DECLINE_MIN:
        stage = "declining"
        roi = "moderate"
    else:
        stage = "mid_cycle"
        roi = "good"

    if decline_ratio >= 0.4:
        decline_risk = "high"
    elif decline_ratio >= 0.2:
        decline_risk = "medium"
    else:
        decline_risk = "low"

    return {
        "estimated_stage": stage,
        "aging_farms": aging_count,
        "decline_risk_level": decline_risk,
        "replanting_need_ratio": round(decline_ratio, 4),
        "long_term_roi_potential": roi,
    }


def _detect_numeric_anomalies(cleaned_records: list[dict[str, Any]], target_types: dict[str, str]) -> list[dict[str, Any]]:
    anomalies = []
    for column, col_type in target_types.items():
        if col_type not in {"integer", "float"}:
            continue
        values = [_to_float(row.get(column)) for row in cleaned_records]
        values = [value for value in values if value is not None]
        if len(values) < 8:
            continue

        mean_value = sum(values) / len(values)
        variance = sum((value - mean_value) ** 2 for value in values) / len(values)
        std_dev = math.sqrt(variance)
        if std_dev == 0:
            continue

        outliers = [value for value in values if abs((value - mean_value) / std_dev) >= 2.8]
        if outliers:
            anomalies.append(
                {
                    "column": column,
                    "outlier_count": len(outliers),
                    "sample_outliers": [round(v, 4) for v in outliers[:5]],
                    "mean": round(mean_value, 4),
                    "std_dev": round(std_dev, 4),
                }
            )
    return anomalies


def _visualization_recommendations(
    schema: list[dict[str, Any]],
    geospatial_fields: list[str],
    lifecycle_indicators: list[str],
) -> list[dict[str, str]]:
    numeric_cols = [col["column"] for col in schema if col.get("type") in {"integer", "float"}]
    categorical_cols = [col["column"] for col in schema if col.get("type") in {"string", "boolean"}]
    temporal_cols = [col["column"] for col in schema if col.get("type") == "date"]

    recommendations = [
        {
            "chart": "kpi_cards",
            "reason": "Executive monitoring for portfolio totals and risk signals.",
        },
        {
            "chart": "pivot_table",
            "reason": "Drill-through analytics by region, district, and farm.",
        },
    ]

    if numeric_cols and categorical_cols:
        recommendations.append({
            "chart": "bar_chart",
            "reason": f"Compare {numeric_cols[0]} across {categorical_cols[0]} segments.",
        })

    if len(numeric_cols) >= 2:
        recommendations.append({
            "chart": "scatter_plot",
            "reason": f"Detect correlation between {numeric_cols[0]} and {numeric_cols[1]}.",
        })
        recommendations.append({
            "chart": "correlation_matrix",
            "reason": "Surface multivariate numeric dependencies and risk clusters.",
        })

    if temporal_cols and numeric_cols:
        recommendations.append({
            "chart": "line_chart",
            "reason": f"Trend {numeric_cols[0]} over {temporal_cols[0]}.",
        })
        recommendations.append({
            "chart": "timeline_chart",
            "reason": "Track farm lifecycle and operational changes over time.",
        })

    if geospatial_fields:
        recommendations.append({
            "chart": "maps",
            "reason": "Visualize farm clusters, risk hotspots, and ESG overlays.",
        })
        recommendations.append({
            "chart": "heatmap",
            "reason": "Highlight geographic concentrations for risk and productivity.",
        })

    if lifecycle_indicators:
        recommendations.append({
            "chart": "decomposition_tree",
            "reason": "Break down cocoa lifecycle impact on ROI and sustainability.",
        })

    return recommendations[:12]


def _score_band(value: float) -> str:
    if value >= 75:
        return "high"
    if value >= 50:
        return "moderate"
    return "low"


def _build_ai_investment_summary(
    cleaned_records: list[dict[str, Any]],
    schema: list[dict[str, Any]],
    target_types: dict[str, str],
    duplicate_rows: int,
    rows_with_missing_values: int,
    lifecycle: dict[str, Any],
    anomaly_report: list[dict[str, Any]],
) -> dict[str, Any]:
    row_count = len(cleaned_records)
    numeric_columns = [item["column"] for item in schema if item.get("type") in {"integer", "float"}]
    date_columns = [item["column"] for item in schema if item.get("type") == "date"]

    field_size_col = _find_column(numeric_columns, ("field_size_acres", "size", "hectares", "area"))
    yield_col = _find_column(numeric_columns, ("yield_estimate", "yield", "production", "tons"))
    risk_proxy_col = _find_column(numeric_columns, ("risk", "loss", "disease", "exposure"))

    avg_size = None
    avg_yield = None
    if field_size_col:
        values = [_to_float(row.get(field_size_col)) for row in cleaned_records]
        values = [v for v in values if v is not None]
        avg_size = round(sum(values) / len(values), 2) if values else None
    if yield_col:
        values = [_to_float(row.get(yield_col)) for row in cleaned_records]
        values = [v for v in values if v is not None]
        avg_yield = round(sum(values) / len(values), 2) if values else None

    missing_ratio = round(rows_with_missing_values / row_count, 4) if row_count else 0
    anomaly_rows = sum(item.get("outlier_count", 0) for item in anomaly_report)
    anomaly_ratio = round(anomaly_rows / max(row_count, 1), 4)
    duplicate_ratio = round(duplicate_rows / max(row_count + duplicate_rows, 1), 4)

    productivity_score = 70
    if avg_yield is not None:
        productivity_score = min(95, max(35, int(45 + avg_yield * 0.12)))
    productivity_score -= int(missing_ratio * 20)
    productivity_score -= int(anomaly_ratio * 15)
    productivity_score = max(10, min(100, productivity_score))

    sustainability_score = 68
    if _find_column([item["column"] for item in schema], ("agroforestry", "shade", "carbon", "irrigation", "soil")):
        sustainability_score += 8
    sustainability_score -= int(anomaly_ratio * 10)
    sustainability_score = max(10, min(100, sustainability_score))

    risk_score = 30
    risk_score += int(missing_ratio * 30)
    risk_score += int(duplicate_ratio * 15)
    risk_score += int(anomaly_ratio * 20)
    if lifecycle.get("decline_risk_level") == "high":
        risk_score += 18
    elif lifecycle.get("decline_risk_level") == "medium":
        risk_score += 9
    if risk_proxy_col:
        risk_score += 5
    risk_score = max(5, min(100, risk_score))

    investment_potential_score = max(0, min(100, int((productivity_score * 0.45) + (sustainability_score * 0.35) + ((100 - risk_score) * 0.20))))

    confidence = 75
    if row_count < 100:
        confidence -= 12
    if missing_ratio > 0.2:
        confidence -= 15
    if not date_columns:
        confidence -= 8
    confidence = max(35, min(96, confidence))

    key_insights = [
        f"Dataset includes {row_count} clean rows across {len(schema)} columns.",
        f"Cocoa lifecycle stage estimated as {lifecycle.get('estimated_stage', 'unknown').replace('_', ' ')}.",
        f"Detected {len(anomaly_report)} numeric anomaly channels for targeted investigation.",
    ]
    if avg_size is not None:
        key_insights.append(f"Average field size is {avg_size} acres-equivalent.")
    if avg_yield is not None:
        key_insights.append(f"Average yield proxy is {avg_yield} units.")

    risk_flags = []
    if missing_ratio >= 0.15:
        risk_flags.append("Elevated missing data risk may reduce forecast reliability.")
    if anomaly_ratio >= 0.08:
        risk_flags.append("Outlier density indicates potential operational instability.")
    if lifecycle.get("decline_risk_level") in {"high", "medium"}:
        risk_flags.append("Aging cocoa lifecycle profile may affect long-term output.")
    if not risk_flags:
        risk_flags.append("No critical systemic risk flags detected in current dataset sample.")

    recommended_actions = [
        "Prioritize farms with high productivity and medium-to-low lifecycle decline risk.",
        "Launch data quality remediation for rows with missing yield and location fields.",
        "Schedule targeted agronomy interventions where anomaly signals cluster.",
    ]
    if lifecycle.get("replanting_need_ratio", 0) >= 0.2:
        recommended_actions.append("Plan phased replanting over the next 5-8 years for aging farm clusters.")

    advisory_note = (
        "Partner Advisory Note:\n"
        "This farm portfolio demonstrates moderate long-term investment potential. "
        "Several farms are approaching the late economic productivity stage of cocoa production and may require "
        "phased replanting strategies within 5-8 years. Increased agroforestry adoption and irrigation upgrades "
        "could improve sustainability and long-term ROI."
    )

    return {
        "executive_summary": (
            f"Portfolio screening indicates {_score_band(investment_potential_score)} investment potential with "
            f"{_score_band(100 - risk_score)} operational risk exposure and {_score_band(sustainability_score)} sustainability posture."
        ),
        "key_insights": key_insights,
        "risk_flags": risk_flags,
        "recommended_actions": recommended_actions,
        "investment_potential_score": investment_potential_score,
        "sustainability_score": sustainability_score,
        "productivity_score": productivity_score,
        "cocoa_lifecycle_assessment": lifecycle,
        "replanting_recommendation": "Required" if lifecycle.get("replanting_need_ratio", 0) >= 0.2 else "Monitor",
        "confidence_level": confidence,
        "partner_advisory_note": advisory_note,
    }


def parse_and_clean_dataset(file_bytes: bytes, filename: str) -> dict[str, Any]:
    file_format = detect_file_format(filename)
    if file_format == "csv":
        headers, raw_records = _read_csv(file_bytes)
    elif file_format == "xlsx":
        headers, raw_records = _read_xlsx(file_bytes)
    else:
        headers, raw_records = _read_json(file_bytes)

    if not raw_records:
        raise DatasetImportError("No data rows found in the uploaded file.")

    kinds_by_column: dict[str, Counter] = {header: Counter() for header in headers}
    missing_counts: dict[str, int] = {header: 0 for header in headers}

    for record in raw_records:
        for header in headers:
            value = _normalize_missing(record.get(header))
            record[header] = value
            if value is None:
                missing_counts[header] += 1
            kinds_by_column[header][_value_kind(value)] += 1

    schema = []
    target_types: dict[str, str] = {}
    for header in headers:
        target = _final_column_type(kinds_by_column[header])
        target_types[header] = target
        schema.append(
            {
                "column": header,
                "type": target,
                "missing_count": missing_counts[header],
                "missing_ratio": round(missing_counts[header] / len(raw_records), 4),
            }
        )

    cleaned_records: list[dict[str, Any]] = []
    seen = set()
    duplicate_rows = 0
    inconsistent_rows = 0
    rows_with_missing_values = 0
    columnar_samples: dict[str, list[Any]] = {header: [] for header in headers}
    top_categorical_counters: dict[str, Counter] = {
        header: Counter() for header in headers if target_types[header] in {"string", "boolean"}
    }
    aggregate_stats: dict[str, dict[str, Any]] = {
        header: {
            "type": target_types[header],
            "non_null_count": 0,
            "null_count": 0,
            "min": None,
            "max": None,
            "sum": 0.0,
            "mean": None,
            "top_values": [],
        }
        for header in headers
    }

    for record in raw_records:
        cleaned: dict[str, Any] = {}
        row_inconsistent = False
        row_has_missing = False

        for header in headers:
            value, inconsistent = _coerce_value(record.get(header), target_types[header])
            if inconsistent:
                row_inconsistent = True
            if value is None:
                row_has_missing = True
            cleaned[header] = value

        if row_inconsistent:
            inconsistent_rows += 1
        if row_has_missing:
            rows_with_missing_values += 1

        fingerprint = _fingerprint(cleaned)
        if fingerprint in seen:
            duplicate_rows += 1
            continue

        seen.add(fingerprint)
        cleaned_records.append(cleaned)

        for header in headers:
            value = cleaned.get(header)
            target_type = target_types[header]
            column_stats = aggregate_stats[header]

            if value is None:
                column_stats["null_count"] += 1
                if len(columnar_samples[header]) < COLUMNAR_SAMPLE_LIMIT:
                    columnar_samples[header].append(None)
                continue

            column_stats["non_null_count"] += 1
            if len(columnar_samples[header]) < COLUMNAR_SAMPLE_LIMIT:
                columnar_samples[header].append(value)

            if target_type in {"integer", "float"}:
                try:
                    numeric_value = float(value)
                except (TypeError, ValueError):
                    continue
                column_stats["sum"] += numeric_value
                if column_stats["min"] is None or numeric_value < column_stats["min"]:
                    column_stats["min"] = numeric_value
                if column_stats["max"] is None or numeric_value > column_stats["max"]:
                    column_stats["max"] = numeric_value
            elif target_type in {"string", "boolean"}:
                top_categorical_counters[header][str(value)] += 1

    for header in headers:
        column_stats = aggregate_stats[header]
        non_null_count = column_stats["non_null_count"]
        if non_null_count and column_stats["type"] in {"integer", "float"}:
            column_stats["mean"] = round(column_stats["sum"] / non_null_count, 4)
            column_stats["sum"] = round(column_stats["sum"], 4)
            if column_stats["min"] is not None:
                column_stats["min"] = round(float(column_stats["min"]), 4)
            if column_stats["max"] is not None:
                column_stats["max"] = round(float(column_stats["max"]), 4)
        else:
            column_stats["sum"] = None

        if header in top_categorical_counters:
            column_stats["top_values"] = [
                {"label": label, "count": count}
                for label, count in top_categorical_counters[header].most_common(CATEGORICAL_TOP_LIMIT)
            ]

    geospatial_fields = _detect_geospatial_fields(headers)
    lifecycle_indicators = _detect_lifecycle_indicators(headers)
    lifecycle_assessment = _estimate_cocoa_lifecycle(cleaned_records, headers)
    anomaly_report = _detect_numeric_anomalies(cleaned_records, target_types)

    metadata_profile = {
        "schema_auto_detected": True,
        "numeric_fields": [item["column"] for item in schema if item.get("type") in {"integer", "float"}],
        "categorical_fields": [item["column"] for item in schema if item.get("type") in {"string", "boolean"}],
        "temporal_fields": [item["column"] for item in schema if item.get("type") == "date"],
        "geospatial_fields": geospatial_fields,
        "farm_lifecycle_indicators": lifecycle_indicators,
        "missing_value_columns": [
            item["column"] for item in schema if item.get("missing_count", 0) > 0
        ],
        "duplicate_rows_removed": duplicate_rows,
        "anomaly_columns": [item["column"] for item in anomaly_report],
        "lifecycle_assessment": lifecycle_assessment,
    }

    visualization_recommendations = _visualization_recommendations(
        schema=schema,
        geospatial_fields=geospatial_fields,
        lifecycle_indicators=lifecycle_indicators,
    )

    ai_summary = _build_ai_investment_summary(
        cleaned_records=cleaned_records,
        schema=schema,
        target_types=target_types,
        duplicate_rows=duplicate_rows,
        rows_with_missing_values=rows_with_missing_values,
        lifecycle=lifecycle_assessment,
        anomaly_report=anomaly_report,
    )

    return {
        "file_format": file_format,
        "schema": schema,
        "rows": cleaned_records,
        "stats": {
            "total_rows_read": len(raw_records),
            "rows_after_dedup": len(cleaned_records),
            "duplicate_rows_removed": duplicate_rows,
            "rows_with_missing_values": rows_with_missing_values,
            "rows_with_inconsistent_values": inconsistent_rows,
            "aggregated_columns": aggregate_stats,
            "metadata_profile": metadata_profile,
            "anomaly_report": anomaly_report,
            "visualization_recommendations": visualization_recommendations,
            "ai_investment_insight_summary": ai_summary,
            "columnar_storage": {
                "sample_row_limit": COLUMNAR_SAMPLE_LIMIT,
                "sample_row_count": max((len(values) for values in columnar_samples.values()), default=0),
                "columns": {
                    header: {
                        "type": target_types[header],
                        "values": columnar_samples[header],
                    }
                    for header in headers
                },
            },
        },
    }
